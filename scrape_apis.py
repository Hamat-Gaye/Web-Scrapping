import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook, Workbook

PORT = 5500

# Load input Excel file
input_file = 'AmatGayetesteasy.xlsx'
sheet_name = 'service_api'

wb = load_workbook(input_file)
ws = wb[sheet_name]

# Prepare output Excel file
output_wb = Workbook()
output_ws = output_wb.active
output_ws.title = "Results"
output_ws.append(["Service Name", "API Route", "Total", "Failures", "Errors", "Skipped"])

# Iterate through rows reading Service Name and API Route
for row in ws.iter_rows(min_row=2, values_only=True): 
    service_name = row[0]    
    api_route = row[1]     
    
    url = f"http://{api_route.replace('{{PORT}}', str(PORT)+'/apis')}"

    try:
        response = requests.get(url)
        response.raise_for_status()

        soup = BeautifulSoup(response.text, "html.parser")


        table = soup.find("table", id="unit-tests")
        if not table:
            raise ValueError("Missing table")

        headers = [th.text.strip().lower() for th in table.find("tr").find_all("th")]
        values = table.find_all("tr")[1].find_all("td")

        data = {headers[i]: values[i].text.strip() for i in range(len(headers))}

        total = data.get("total", "N/A")
        failures = data.get("failures", "N/A")
        errors = data.get("errors", "N/A")
        skipped = data.get("skipped", "N/A")

    except Exception as e:
        print(f'An error occured {e}')
        total = failures = errors = skipped = "N/A"

    output_ws.append([service_name, url, total, failures, errors, skipped])

output_wb.save("api_unit_test_results.xlsx")
print("Scraping completed. Results saved to api_unit_test_results.xlsx.")
